import re
from functools import partial
import datetime
from docx import Document
from openpyxl import load_workbook


def cleanse_str(s):
    return ''.join(c for c in s if c.isalnum())


def run(data_xlsx_path, template_docx_path, result_folder_path='tests/results'):
    print('\n***** SCRIPT STARTS *****')

    # ========== Initialising data ==========

    wb = load_workbook(data_xlsx_path)
    ws = wb.active  # Get the first worksheet

    repl_data = {}
    first_row = ws[1]
    data_cols = tuple(ws.columns)

    # First row of worksheet contains variable names
    for cell in first_row:
        repl_data[cleanse_str(cell.value)] = []

    # The remaining rows contain variable values
    for col in data_cols:
        for i in range(1, len(col)):
            newStr = '{:%d %b %Y}'.format(col[i].value) if isinstance(col[i].value, datetime.date) else col[i].value
            repl_data[cleanse_str(col[0].value)].append(newStr)

    # ========== Work with .docx ==========

    REGEX_PAT = '\[(.*?)\]'

    def repl_func(matched_obj, r_index):
        if matched_obj:
            for matched_str in matched_obj.groups():
                return repl_data[cleanse_str(matched_str)][r_index]

    for row_index in range(0, ws.max_row - 1):
        document = Document(template_docx_path)
        paragraphs = document.paragraphs
        print('Working on row_index#{}'.format(row_index))
        for para_index, paragraph in enumerate(paragraphs):
            if paragraph.text != '':
                match = re.findall(REGEX_PAT, paragraph.text)
                if match:
                    paragraph.text = re.sub(REGEX_PAT, partial(repl_func, r_index=row_index), paragraph.text)
                    print('Replaced match. New paragraph: {}'.format(paragraph.text))
        document.save('{}/res_{}.docx'.format(result_folder_path, row_index + 1))

    print('\nOperations finished with no error :).')
