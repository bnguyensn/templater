import re
from functools import partial
import datetime
from docx import Document
from openpyxl import load_workbook

wb = load_workbook('invconf_data.xlsx')
ws = wb.active  # Get the first worksheet


def cleanse_str(s):
    return ''.join(c for c in s if c.isalnum())


print('\n***** SCRIPT STARTS *****')

# ========== Initialising data ==========

replData = {}
firstRow = ws[1]
dataCols = tuple(ws.columns)

# First row of worksheet contains variable names
for cell in firstRow:
    replData[cleanse_str(cell.value)] = []

# The remaining rows contain variable values
for col in dataCols:
    for i in range(1, len(col)):
        newStr = '{:%d %b %Y}'.format(col[i].value) if isinstance(col[i].value, datetime.date) else col[i].value
        replData[cleanse_str(col[0].value)].append(newStr)

# ========== Work with .docx ==========

REGEX_PAT = '\[(.*?)\]'


def repl_func(matched_obj, r_index):
    if matched_obj:
        for matched_str in matched_obj.groups():
            return replData[cleanse_str(matched_str)][r_index]


for row_index in range(0, ws.max_row - 1):
    document = Document('invconf_template.docx')
    paragraphs = document.paragraphs
    print('Working on row_index#{}'.format(row_index))
    for para_index, paragraph in enumerate(paragraphs):
        if paragraph.text != '':
            match = re.findall(REGEX_PAT, paragraph.text)
            if match:
                paragraph.text = re.sub(REGEX_PAT, partial(repl_func, r_index=row_index), paragraph.text)
                print('Replaced match. New paragraph: {}'.format(paragraph.text))
    document.save('./docx/invconf_#{}.docx'.format(row_index + 1))

# for i, row in enumerate(dataRows):
#     if i > 0:
#         for cell in row:
#             print('\nrow #{}, column #{}, value = {}'.format(cell.row, cell.column, cell.value))

# Replacing text

# print("\nDocument's paragraphs:")
# for i, paragraph in enumerate(paragraphs):
#     if paragraph.text != '':
#         match = re.findall(rePattern, paragraph.text)
#         if match:
#             print('\nFound match at paragraph #{}: {}'.format(i, match))
#             paragraph.text = re.sub(rePattern, 'REDACTED', paragraph.text)
#             print('Replaced match. New paragraph: {}'.format(paragraph.text))

print('\nOperations finished with no error :).')
